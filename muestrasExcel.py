"""
Generador de Archivo Excel para Simulación de Muestreo LogiCorp
Probabilidad y Estadística - Unidad III
Licenciatura en Logística Empresarial

Autor: [Tu nombre]
Fecha: Marzo 2025
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import random
from openpyxl.cell.cell import Cell

def set_formula(ws, cell_ref, formula):
    cell = ws[cell_ref]
    cell.value = formula
    cell.data_type = 'f'

def generar_archivo_logicorp(num_equipos=11, nombre_archivo="LogiCorp_Muestreo.xlsx"):
    """
    Genera archivo Excel completo para la actividad de simulación de muestreo.
    
    Parameters:
    num_equipos (int): Número de equipos (default: 11)
    nombre_archivo (str): Nombre del archivo a generar
    """
    
    # Configurar semilla para reproducibilidad (opcional)
    np.random.seed(42)
    
    # Parámetros poblacionales
    mu_poblacional = 2.35  # Media poblacional
    sigma_poblacional = 0.38  # Desviación estándar poblacional
    tamano_poblacion = 1000
    tamano_muestra = 30
    
    print(f"Generando población de {tamano_poblacion} entregas...")
    print(f"Media poblacional: {mu_poblacional} horas")
    print(f"Desviación estándar: {sigma_poblacional} horas")
    
    # Generar población con distribución normal
    poblacion = np.random.normal(mu_poblacional, sigma_poblacional, tamano_poblacion)
    
    # Asegurar que todos los valores sean positivos y realistas (entre 1 y 4 horas)
    poblacion = np.clip(poblacion, 1.0, 4.0)
    poblacion = np.round(poblacion, 2)
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Estilos para formateo
    font_titulo = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    font_subtitulo = Font(name='Arial', size=12, bold=True, color='003366')
    font_normal = Font(name='Arial', size=10)
    font_formula = Font(name='Arial', size=10, italic=True)
    
    fill_header = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
    fill_subtitle = PatternFill(start_color='E6F2FF', end_color='E6F2FF', fill_type='solid')
    fill_data = PatternFill(start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')
    
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_left = Alignment(horizontal='left', vertical='center')
    
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # CREAR HOJA DE POBLACIÓN
    print("Creando hoja de población...")
    ws_poblacion = wb.create_sheet("POBLACIÓN")
    
    # Headers de población
    ws_poblacion['A1'] = "ID_Entrega"
    ws_poblacion['B1'] = "Tiempo_Entrega"
    ws_poblacion['C1'] = "Descripción"
    
    # Aplicar formato a headers
    for col in ['A1', 'B1', 'C1']:
        ws_poblacion[col].font = font_subtitulo
        ws_poblacion[col].fill = fill_subtitle
        ws_poblacion[col].alignment = alignment_center
        ws_poblacion[col].border = border_thin
    
    # Llenar datos de población
    for i in range(tamano_poblacion):
        ws_poblacion[f'A{i+2}'] = i + 1
        ws_poblacion[f'B{i+2}'] = poblacion[i]
        ws_poblacion[f'C{i+2}'] = f"Entrega {i+1}"
        
        # Formato para datos
        for col in ['A', 'B', 'C']:
            ws_poblacion[f'{col}{i+2}'].font = font_normal
            ws_poblacion[f'{col}{i+2}'].border = border_thin
    
    # Ajustar ancho de columnas
    ws_poblacion.column_dimensions['A'].width = 12
    ws_poblacion.column_dimensions['B'].width = 15
    ws_poblacion.column_dimensions['C'].width = 15
    
    # CREAR HOJAS PARA CADA EQUIPO
    for equipo in range(1, num_equipos + 1):
        print(f"Creando hoja para Equipo {equipo}...")
        
        ws = wb.create_sheet(f"EQUIPO_{equipo}")
        
        # SECCIÓN 1: INFORMACIÓN DEL EQUIPO
        ws['A1'] = f"EQUIPO {equipo} - SIMULACIÓN DE MUESTREO LOGICORP"
        ws['A1'].font = font_titulo
        ws['A1'].fill = fill_header
        ws['A1'].alignment = alignment_center
        ws.merge_cells('A1:F1')
        
        # SECCIÓN 2: PARÁMETROS POBLACIONALES
        ws['A3'] = "PARÁMETROS POBLACIONALES"
        ws['A3'].font = font_subtitulo
        ws['A3'].fill = fill_subtitle
        ws.merge_cells('A3:D3')
        
        ws['A4'] = "Media poblacional (μ):"
        set_formula(ws, 'B4', "=PROMEDIO('POBLACIÓN'!B:B)")
        #ws['B4'] = "=PROMEDIO('POBLACIÓN'!B:B)"
        ws['C4'] = "horas"

        ws['A5'] = "Desv. estándar poblacional (σ):"
        set_formula(ws, 'B5', "=DESVEST.P('POBLACIÓN'!B:B)")
        #ws['B5'] = "=DESVEST.P('POBLACIÓN'!B:B)"
        ws['C5'] = "horas"

        ws['A6'] = "Tamaño de población (N):"
        set_formula(ws, 'B6', "=CONTAR('POBLACIÓN'!B:B)-1")
        #ws['B6'] = "=CONTAR('POBLACIÓN'!B:B)-1"
        ws['C6'] = "entregas"

        
        # Formato para parámetros
        for row in range(4, 7):
            ws[f'A{row}'].font = font_normal
            ws[f'B{row}'].font = font_formula
            ws[f'C{row}'].font = font_normal
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].border = border_thin
        
        # SECCIÓN 3: INSTRUCCIONES
        ws['A8'] = "INSTRUCCIONES:"
        ws['A8'].font = font_subtitulo
        ws['A9'] = "1. Presione F9 para generar nuevas muestras aleatorias"
        ws['A10'] = "2. Los cálculos se actualizan automáticamente"
        ws['A11'] = "3. Compare sus resultados con otros equipos"
        
        for row in range(8, 12):
            ws[f'A{row}'].font = font_normal
        
        # FUNCIÓN PARA CREAR SECCIÓN DE MUESTRA
        def crear_seccion_muestra(ws, num_muestra, fila_inicio):
            # Título de la muestra
            ws[f'A{fila_inicio}'] = f"MUESTRA {num_muestra} (n={tamano_muestra})"
            ws[f'A{fila_inicio}'].font = font_subtitulo
            ws[f'A{fila_inicio}'].fill = fill_subtitle
            ws.merge_cells(f'A{fila_inicio}:D{fila_inicio}')
            
            # Headers de la muestra
            ws[f'A{fila_inicio+1}'] = "ID_Seleccionado"
            ws[f'B{fila_inicio+1}'] = "Tiempo_Entrega"
            
            # Formato headers
            for col in ['A', 'B']:
                ws[f'{col}{fila_inicio+1}'].font = font_subtitulo
                ws[f'{col}{fila_inicio+1}'].fill = fill_data
                ws[f'{col}{fila_inicio+1}'].alignment = alignment_center
                ws[f'{col}{fila_inicio+1}'].border = border_thin
            
            # Fórmulas para generar muestra aleatoria
            for i in range(tamano_muestra):
                fila_actual = fila_inicio + 2 + i
                # ID aleatorio entre 1 y 1000
                ws[f'A{fila_actual}'] = f"=INT(RAND()*1000)+1"
                # Buscar tiempo correspondiente
                ws[f'B{fila_actual}'] = f"=VLOOKUP(A{fila_actual},'POBLACIÓN'!A:B,2,FALSE)"
                # ws[f'B{fila_actual}'] = f"=VLOOKUP(A{fila_actual},POBLACIÓN.A:B,2,FALSE)"
                
                # Formato
                for col in ['A', 'B']:
                    ws[f'{col}{fila_actual}'].font = font_normal
                    ws[f'{col}{fila_actual}'].border = border_thin
            
            # Estadísticos de la muestra
            fila_stats = fila_inicio + 2 + tamano_muestra + 1
            
            ws[f'D{fila_inicio+1}'] = f"ESTADÍSTICOS MUESTRA {num_muestra}:"
            ws[f'D{fila_inicio+1}'].font = font_subtitulo
            ws[f'D{fila_inicio+1}'].fill = fill_subtitle
            
            ws[f'D{fila_inicio+2}'] = f"Media muestral (x̄{num_muestra}):"
            set_formula(ws, f'E{fila_inicio+2}', f"=AVERAGE(B{fila_inicio+2}:B{fila_inicio+1+tamano_muestra})")
            #ws[f'E{fila_inicio+2}'] = f"=AVERAGE(B{fila_inicio+2}:B{fila_inicio+1+tamano_muestra})"
            ws[f'F{fila_inicio+2}'] = "horas"
            
            ws[f'D{fila_inicio+3}'] = f"Desv. muestral (s{num_muestra}):"
            set_formula(ws, f'E{fila_inicio+3}', f"=STDEV.S(B{fila_inicio+2}:B{fila_inicio+1+tamano_muestra})")
            #ws[f'E{fila_inicio+3}'] = f"=STDEV.S(B{fila_inicio+2}:B{fila_inicio+1+tamano_muestra})"
            ws[f'F{fila_inicio+3}'] = "horas"
            
            ws[f'D{fila_inicio+4}'] = f"Error vs μ:"
            set_formula(ws, f'E{fila_inicio+4}', f"=ABS(E{fila_inicio+2}-$B$4)")
            #ws[f'E{fila_inicio+4}'] = f"=ABS(E{fila_inicio+2}-$B$4)"
            ws[f'F{fila_inicio+4}'] = "horas"
            
            # Formato estadísticos
            for row in range(fila_inicio+2, fila_inicio+5):
                ws[f'D{row}'].font = font_normal
                ws[f'E{row}'].font = font_formula
                ws[f'F{row}'].font = font_normal
                for col in ['D', 'E', 'F']:
                    ws[f'{col}{row}'].border = border_thin
            
            return fila_inicio + 2 + tamano_muestra + 5
        
        # Crear las tres secciones de muestras
        fila_actual = 13
        referencias_medias = []
        
        for num_muestra in range(1, 4):
            fila_actual = crear_seccion_muestra(ws, num_muestra, fila_actual)
            # Guardar referencia a la celda de la media para el resumen
            referencias_medias.append(f"E{fila_actual - 3}")
            fila_actual += 2  # Espacio entre muestras
        
        # SECCIÓN 4: RESUMEN DEL EQUIPO
        fila_resumen = fila_actual + 2
        
        ws[f'A{fila_resumen}'] = "RESUMEN DEL EQUIPO"
        ws[f'A{fila_resumen}'].font = font_titulo
        ws[f'A{fila_resumen}'].fill = fill_header
        ws[f'A{fila_resumen}'].alignment = alignment_center
        ws.merge_cells(f'A{fila_resumen}:D{fila_resumen}')
        
        # Tabla resumen
        headers_resumen = ['Muestra', 'Media (x̄)', 'Error vs μ', 'Interpretación']
        for i, header in enumerate(headers_resumen):
            ws[f'{chr(65+i)}{fila_resumen+1}'] = header
            ws[f'{chr(65+i)}{fila_resumen+1}'].font = font_subtitulo
            ws[f'{chr(65+i)}{fila_resumen+1}'].fill = fill_subtitle
            ws[f'{chr(65+i)}{fila_resumen+1}'].alignment = alignment_center
            ws[f'{chr(65+i)}{fila_resumen+1}'].border = border_thin
        
        # Datos del resumen
        for i in range(3):
            fila_dato = fila_resumen + 2 + i
            ws[f'A{fila_dato}'] = i + 1
            ws[f'B{fila_dato}'] = f"={referencias_medias[i]}"
            ws[f'C{fila_dato}'] = f"=ABS(B{fila_dato}-$B$4)"
            ws[f'D{fila_dato}'] = f'=IF(C{fila_dato}<0.1,"Excelente",IF(C{fila_dato}<0.2,"Bueno","Aceptable"))'
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{fila_dato}'].font = font_normal
                ws[f'{col}{fila_dato}'].border = border_thin
        
        # Promedio general
        fila_promedio = fila_resumen + 5
        ws[f'A{fila_promedio}'] = "PROMEDIO:"
        ws[f'B{fila_promedio}'] = f"=AVERAGE(B{fila_resumen+2}:B{fila_resumen+4})"
        ws[f'C{fila_promedio}'] = f"=AVERAGE(C{fila_resumen+2}:C{fila_resumen+4})"
        ws[f'D{fila_promedio}'] = "Resultado final"
        
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{fila_promedio}'].font = font_subtitulo
            ws[f'{col}{fila_promedio}'].fill = fill_subtitle
            ws[f'{col}{fila_promedio}'].border = border_thin
        
        # SECCIÓN 5: EXPERIMENTOS ADICIONALES
        fila_experimentos = fila_promedio + 3
        
        ws[f'A{fila_experimentos}'] = "EXPERIMENTOS ADICIONALES"
        ws[f'A{fila_experimentos}'].font = font_subtitulo
        ws[f'A{fila_experimentos}'].fill = fill_subtitle
        ws.merge_cells(f'A{fila_experimentos}:D{fila_experimentos}')
        
        ws[f'A{fila_experimentos+1}'] = "Presione F9 múltiples veces y anote las medias:"
        ws[f'A{fila_experimentos+2}'] = "Experimento 1:"
        ws[f'B{fila_experimentos+2}'] = "_____"
        ws[f'A{fila_experimentos+3}'] = "Experimento 2:"
        ws[f'B{fila_experimentos+3}'] = "_____"
        ws[f'A{fila_experimentos+4}'] = "Experimento 3:"
        ws[f'B{fila_experimentos+4}'] = "_____"
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10
    
    # CREAR HOJA DE CONSOLIDACIÓN
    print("Creando hoja de consolidación...")
    ws_consolidacion = wb.create_sheet("CONSOLIDACIÓN")
    
    # Título
    ws_consolidacion['A1'] = "CONSOLIDACIÓN DE RESULTADOS - TODOS LOS EQUIPOS"
    ws_consolidacion['A1'].font = font_titulo
    ws_consolidacion['A1'].fill = fill_header
    ws_consolidacion['A1'].alignment = alignment_center
    ws_consolidacion.merge_cells('A1:F1')
    
    # Headers
    headers_consolidacion = ['Equipo', 'Media 1', 'Media 2', 'Media 3', 'Promedio', 'Error Promedio']
    for i, header in enumerate(headers_consolidacion):
        ws_consolidacion[f'{chr(65+i)}3'] = header
        ws_consolidacion[f'{chr(65+i)}3'].font = font_subtitulo
        ws_consolidacion[f'{chr(65+i)}3'].fill = fill_subtitle
        ws_consolidacion[f'{chr(65+i)}3'].alignment = alignment_center
        ws_consolidacion[f'{chr(65+i)}3'].border = border_thin
    
    # Referencias a equipos
    for equipo in range(1, num_equipos + 1):
        fila = 3 + equipo
        ws_consolidacion[f'A{fila}'] = equipo
        # Las fórmulas dependerán de la estructura exacta creada arriba
        ws_consolidacion[f'B{fila}'] = f"=EQUIPO_{equipo}.E14"  # Ajustar según ubicación real
        ws_consolidacion[f'C{fila}'] = f"=EQUIPO_{equipo}.E45"  # Ajustar según ubicación real
        ws_consolidacion[f'D{fila}'] = f"=EQUIPO_{equipo}.E76"  # Ajustar según ubicación real
        ws_consolidacion[f'E{fila}'] = f"=AVERAGE(B{fila}:D{fila})"
        ws_consolidacion[f'F{fila}'] = f"=ABS(E{fila}-'POBLACIÓN'!B2)"
        # ws_consolidacion[f'F{fila}'] = f"=ABS(E{fila}-POBLACIÓN.B2)"  # Error vs media poblacional
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_consolidacion[f'{col}{fila}'].font = font_normal
            ws_consolidacion[f'{col}{fila}'].border = border_thin
    
    # Estadísticos generales
    fila_stats_generales = 3 + num_equipos + 2
    ws_consolidacion[f'A{fila_stats_generales}'] = "ESTADÍSTICOS GENERALES"
    ws_consolidacion[f'A{fila_stats_generales}'].font = font_subtitulo
    ws_consolidacion[f'A{fila_stats_generales}'].fill = fill_subtitle
    
    ws_consolidacion[f'A{fila_stats_generales+1}'] = "Media de todas las medias:"
    ws_consolidacion[f'B{fila_stats_generales+1}'] = f"=AVERAGE(E4:E{3+num_equipos})"
    
    ws_consolidacion[f'A{fila_stats_generales+2}'] = "Desv. est. de las medias:"
    ws_consolidacion[f'B{fila_stats_generales+2}'] = f"=STDEV.S(E4:E{3+num_equipos})"
    
    ws_consolidacion[f'A{fila_stats_generales+3}'] = "Error estándar teórico:"
    ws_consolidacion[f'B{fila_stats_generales+3}'] = f"=POBLACIÓN.B2/SQRT(30)"
    
    # Ajustar columnas
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_consolidacion.column_dimensions[col].width = 15
    
    # Guardar archivo
    print(f"Guardando archivo: {nombre_archivo}")
    wb.save(nombre_archivo)
    
    # Estadísticas finales
    media_real = np.mean(poblacion)
    std_real = np.std(poblacion, ddof=0)
    
    print(f"\n📊 ARCHIVO GENERADO EXITOSAMENTE!")
    print(f"📁 Nombre: {nombre_archivo}")
    print(f"📈 Estadísticas reales de la población generada:")
    print(f"   • Media real: {media_real:.3f} horas")
    print(f"   • Desviación estándar real: {std_real:.3f} horas")
    print(f"   • Rango: {np.min(poblacion):.2f} - {np.max(poblacion):.2f} horas")
    print(f"🔢 Hojas creadas: {num_equipos + 2} (1 población + {num_equipos} equipos + 1 consolidación)")
    
    return wb

if __name__ == "__main__":
    # Configuración
    NUM_EQUIPOS = 11
    NOMBRE_ARCHIVO = "LogiCorp_Muestreo.xlsx"
    
    print("🚀 Iniciando generación del archivo Excel para LogiCorp...")
    print("=" * 60)
    
    try:
        workbook = generar_archivo_logicorp(NUM_EQUIPOS, NOMBRE_ARCHIVO)
        print("=" * 60)
        print("✅ ¡Archivo generado correctamente!")
        print(f"📋 Listo para usar en clase con {NUM_EQUIPOS} equipos")
        print("\n🔧 Para usar el archivo:")
        print("1. Abrir en Excel")
        print("2. Cada equipo va a su pestaña correspondiente")
        print("3. Presionar F9 para generar muestras aleatorias")
        print("4. Observar los resultados automáticos")
        
    except Exception as e:
        print(f"❌ Error al generar el archivo: {e}")
        print("📝 Verificar que tengas instalados: pandas, numpy, openpyxl")
        print("💿 Instalar con: pip install pandas numpy openpyxl")